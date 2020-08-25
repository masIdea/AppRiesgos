from django.shortcuts import render
from datetime import datetime
import xlsxwriter
from appCarga.clases.appCarga.Carga import XlsxWriterObj
from appCarga.clases.appCarga.ModelosCargaDatos import ModelosCargaDatos
from appCarga.clases.appCarga.Auditoria import Auditoria
from core.models import RiesgoListas, RiesgoCargasistemacio, RiesgoAuditoria, RiesgoCargasisteman1, RiesgoRbsfamilia, RiesgoListas, RiesgoUnifica
import openpyxl 
from core.clases.core.identificadores import GetId
import io
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from crearRiesgo.queries import QueryRiesgoSimilar, QueryFamilia, QuerySubproceso
from core.clases.core.bulk import BulkCreateManager
from core.utils import id_generator
from django.contrib.auth.decorators import login_required

# Create your views here.
@login_required
def inicio(request):
    if 'readcreate' in request.session:        
        return render(request, "appCarga/inicio.html")
    else:
        return HttpResponseRedirect(reverse('inicio_dashboard'))

@login_required
def datosCIO(request):
    familias = list(QueryFamilia())
    subprocesos = list(RiesgoRbsfamilia.objects.values("subproceso").distinct())
    clasificacion = list(RiesgoListas.objects.filter(tipo="Clasificaciones").values())
    AreaDeEvento = list(RiesgoListas.objects.filter(tipo="AreaDeEvento").values())
    SuperIntendenciaEvento = list(RiesgoListas.objects.filter(tipo="SuperIntendenciaEvento").values())
    Proyecto = list(RiesgoListas.objects.filter(tipo="Proyecto").values())    
    ControlesQueNoFuncionaron = list(RiesgoListas.objects.filter(tipo="ControlesQueNoFuncionaron").values())
    TipoFalla = list(RiesgoListas.objects.filter(tipo="TipoFalla").values())

    data = {
            'familias':familias, 
            'subprocesos':subprocesos,
            'clasificacion':clasificacion,
            'AreaDeEvento':AreaDeEvento,
            'SuperIntendenciaEvento':SuperIntendenciaEvento,
            'Proyecto':Proyecto,            
            'ControlesQueNoFuncionaron':ControlesQueNoFuncionaron,
            'TipoFalla':TipoFalla,
            }
    if 'readcreate' in request.session:
        return render(request, "appCarga/datosCio.html", data)
    else:
        return HttpResponseRedirect(reverse('inicio_dashboard'))

def getDatosCIO(request):
    if request.method == "GET":
        datos = list(RiesgoCargasistemacio.objects.all().values())
        data = {'datos':datos}
        return JsonResponse(data, safe=False)

def getDatosN1(request):
    if request.method == "GET":
        datos = list(RiesgoCargasisteman1.objects.all().values())
        data = {'datos':datos}
        return JsonResponse(data, safe=False)

def getDatosCIOId(request):
    if request.method == "GET":
        id_registro = request.GET['id']
        print(id_registro)
        datos = list(RiesgoCargasistemacio.objects.filter(idcio=id_registro).values())        
        data = {'datos':datos}
        return JsonResponse(data, safe=False)

def DashboardIdDashboard(request):
    if request.method == "GET":
        id_registro = request.GET['id']        
        datos = list(RiesgoUnifica.objects.filter(idregistro=id_registro).values())        
        data = {'datos':datos}
        return JsonResponse(data, safe=False)


def getDatosN1Id(request):
    if request.method == "GET":
        id_registro = request.GET['id']
        datos = list(RiesgoCargasisteman1.objects.filter(idn1=id_registro).values())
        print(datos)
        data = {'datos':datos}
        return JsonResponse(data, safe=False)

def editarCIO(request):
    if request.method == "POST":
        id_cio = request.POST.get("txt-cio")
        IdCosto = request.POST.get("IdCosto")
        Fecha = request.POST.get("Fecha")
        Gerencia = request.POST.get("Gerencia")
        SuperIntendencia = request.POST.get("SuperIntendencia")
        UnidadOperacional = request.POST.get("UnidadOperacional")
        TipoObjeto = request.POST.get("TipoObjeto")
        Evento = request.POST.get("Evento")
        NivelImpacto = request.POST.get("NivelImpacto")
        Comentarios = request.POST.get("Comentarios")
        CantidadDiasDetencion = request.POST.get("CantidadDiasDetencion")
        TipoFalla = request.POST.get("TipoFalla")
        CantidadTMFPerdidas = request.POST.get("CantidadTMFPerdidas")
        MontoDePerdida = request.POST.get("MontoDePerdida")
        RepeticionDelEvento = request.POST.get("RepeticionDelEvento")
        Estadistica = request.POST.get("Estadistica")
        AreaDeEvento = request.POST.get("AreaDeEvento")
        SuperIntendenciaEvento = request.POST.get("SuperIntendenciaEvento")
        Proyecto = request.POST.get("Proyecto")
        Familia = request.POST.get("Familia")
        SubProceso = request.POST.get("SubProceso")
        Clasificacion = request.POST.get("Clasificacion")
        ControlesQueNoFuncionaron = request.POST.get("ControlesQueNoFuncionaron")
        RiesgoCargasistemacio.objects.filter(idcio=id_cio).update(
                                                            idcosto= IdCosto,
                                                            fecha= Fecha,
                                                            gerencia= Gerencia,
                                                            superintendencia= SuperIntendencia,
                                                            unidadoperacional= UnidadOperacional,
                                                            tipoobjeto= TipoObjeto,
                                                            evento= Evento,
                                                            nivelimpacto= NivelImpacto,
                                                            comentarios= Comentarios,
                                                            cantidaddiasdetencion= CantidadDiasDetencion,
                                                            cantidadtmfperdidas= CantidadTMFPerdidas,
                                                            montodeperdida= MontoDePerdida,
                                                            repeticiondelevento= RepeticionDelEvento,
                                                            estadistica= Estadistica,
                                                            areadeevento= AreaDeEvento,
                                                            superintendenciaevento= SuperIntendenciaEvento,
                                                            proyecto= Proyecto,
                                                            familia= Familia,
                                                            subproceso= SubProceso,
                                                            clasificacion= Clasificacion,
                                                            controlesquenofuncionaron= ControlesQueNoFuncionaron,
                                                            tipofalla= TipoFalla,   
                                                            fechamodificacion=datetime.now()                                                                 
                                                                )
        return HttpResponseRedirect(reverse('datos_cio'))

def editarN1(request):
    if request.method == "POST":
        id_n1 = request.POST.get('txt-n1')
        codigorsso = request.POST.get('CodigoRSSO')
        tiporsso = request.POST.get('TipoRSSO')
        estado = request.POST.get('Estado')
        causal = request.POST.get('Causal')
        nivelrsso = request.POST.get('NivelRSSO')
        fechacreacion = request.POST.get('FechaCreacion')
        fechahallazgo = request.POST.get('FechaHallazgo')
        division = request.POST.get('Division')
        gerencia = request.POST.get('Gerencia')
        superintendencia = request.POST.get('Superintendencia')
        area = request.POST.get('Area')
        sapinformante = request.POST.get('SAPInformante')
        informante = request.POST.get('Informante')
        sapresponsable = request.POST.get('SAPResponsable')
        responsable = request.POST.get('Responsable')
        fechacierre = request.POST.get('FechaCierre')
        descripcionincidente = request.POST.get('DescripcionIncidente')
        accionrealizada = request.POST.get('AccionRealizada')
        estandar = request.POST.get('Estandar')
        riesgocritico = request.POST.get('RiesgoCritico')
        idevento = request.POST.get('IdEvento')
        cantidaddiasdetencion = request.POST.get('CantidadDiasDetencion')
        cantidadtmfperdidas = request.POST.get('CantidadTMFPerdidas')
        areadeevento = request.POST.get('AreaDeEvento')
        superintendenciaevento = request.POST.get('SuperIntendenciaEvento')
        proyecto = request.POST.get('Proyecto')
        familia = request.POST.get('Familia')
        subproceso = request.POST.get('SubProceso')
        clasificacion = request.POST.get('Clasificacion')
        controlesquenofuncionaron = request.POST.get('ControlesQueNoFuncionaron')
        montodeperdida = request.POST.get('MontoDePerdida')
        repeticiondelevento = request.POST.get('RepeticionDelEvento')
        tipofalla = request.POST.get('TipoFalla')
        estadistica = request.POST.get('Estadistica')
        RiesgoCargasisteman1.objects.filter(idn1=id_n1).update(
                                                            codigorsso = codigorsso,
                                                            tiporsso = tiporsso,
                                                            estado = estado,
                                                            causal = causal,
                                                            nivelrsso = nivelrsso,
                                                            fechacreacion = fechacreacion,
                                                            fechahallazgo = fechahallazgo,
                                                            division = division,
                                                            gerencia = gerencia,
                                                            superintendencia = superintendencia,
                                                            area = area,
                                                            sapinformante = sapinformante,
                                                            informante = informante,
                                                            sapresponsable = sapresponsable,
                                                            responsable = responsable,
                                                            fechacierre = fechacierre,
                                                            descripcionincidente = descripcionincidente,
                                                            accionrealizada = accionrealizada,
                                                            estandar = estandar,
                                                            riesgocritico = riesgocritico,
                                                            idevento = idevento,
                                                            cantidaddiasdetencion = cantidaddiasdetencion,
                                                            cantidadtmfperdidas = cantidadtmfperdidas,
                                                            areadeevento = areadeevento,
                                                            superintendenciaevento = superintendenciaevento,
                                                            proyecto = proyecto,
                                                            familia = familia,
                                                            subproceso = subproceso,
                                                            clasificacion = clasificacion,
                                                            controlesquenofuncionaron = controlesquenofuncionaron,
                                                            montodeperdida = montodeperdida,
                                                            repeticiondelevento = repeticiondelevento,
                                                            tipofalla = tipofalla,
                                                            estadistica = estadistica,
                                                            fechamodificacion=datetime.now()
                                                                )
        return HttpResponseRedirect(reverse('datos_n1'))

@login_required
def datosN1(request):
    familias = list(QueryFamilia())
    subprocesos = list(RiesgoRbsfamilia.objects.values("subproceso").distinct())
    clasificacion = list(RiesgoListas.objects.filter(tipo="Clasificaciones").values())
    AreaDeEvento = list(RiesgoListas.objects.filter(tipo="AreaDeEvento").values())
    SuperIntendenciaEvento = list(RiesgoListas.objects.filter(tipo="SuperIntendenciaEvento").values())
    Proyecto = list(RiesgoListas.objects.filter(tipo="Proyecto").values())    
    ControlesQueNoFuncionaron = list(RiesgoListas.objects.filter(tipo="ControlesQueNoFuncionaron").values())
    TipoFalla = list(RiesgoListas.objects.filter(tipo="TipoFalla").values())

    data = {
            'familias':familias, 
            'subprocesos':subprocesos,
            'clasificacion':clasificacion,
            'AreaDeEvento':AreaDeEvento,
            'SuperIntendenciaEvento':SuperIntendenciaEvento,
            'Proyecto':Proyecto,            
            'ControlesQueNoFuncionaron':ControlesQueNoFuncionaron,
            'TipoFalla':TipoFalla,
            }
    if 'readcreate' in request.session:            
        return render(request, "appCarga/datosN1.html", data)
    else:
        return HttpResponseRedirect(reverse('inicio_dashboard'))

def generaArchivo(request):        
    archivo_lista = request.POST.getlist("radio-carga-datos")    
    for archivo in archivo_lista:
        if archivo != str(0):
            print("EL ARCHIVO ES ", archivo)
            Auditoria(request.user, "Genera archivo "+str(archivo), "CargaDatos", None).saveAudit()
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})            

            datos = ModelosCargaDatos(archivo).datosColumnasArchivo()
            XlsxWriterObj(datos['arreglo_columnas'], datos['arreglo_listas'], '', '', '', workbook, datos['arreglo_fechas']).closeBook()
            workbook.close()            
            output.seek(0)
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="'+archivo+'.xlsx"'
            return response            

def cargaDatos(request):
    if request.method == "POST":
        archivo = request.POST.get("radio-carga-datos")        
        myfile = request.FILES['archivo']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        path = settings.MEDIA_ROOT+'//'+myfile.name
        print("EL PATH ES ", path)
        wb_obj = openpyxl.load_workbook(path) 
        sheet_obj = wb_obj.active 
        max_col = sheet_obj.max_column
        print("MAXIMO COLUMNAS ", max_col)
        m_row = sheet_obj.max_row


 

        if archivo == "ArchivoCargaSistemaCIO":
            id_carga = GetId('RiesgoCargasistemacio', 'RCARCIO', 'idcargacio').get_id()
            for j in range(2, m_row + 1):
                idcio = GetId('RiesgoCargasistemacio', 'IDCIO', 'idcio').get_id()
                str_query_orm = "RiesgoCargasistemacio( idcio = '"+str(idcio)+"', idcargacio = '"+str(id_carga)+"', estadoregistro='ACTIVO', fecharegistro='"+str(datetime.now())+"', fechamodificacion='"+str(datetime.now())+"',"                
                for i in range(1, max_col + 1):
                    cell_obj_column = sheet_obj.cell(row = 1, column = i)                 
                    cell_obj = sheet_obj.cell(row = j, column = i)
                    datos_celda = cell_obj.value
                    if datos_celda is None:
                        datos_celda = "-"
                    #print(cell_obj_column.value)
                    column = cell_obj_column.value.lower()
                    str_query_orm += str(column)+"='"+str(datos_celda)+"',"
                str_query_orm+=")"
                bulk_mgr = BulkCreateManager()
                bulk_mgr.add(
                    eval(str_query_orm)
                )
                bulk_mgr.done()
                
                #eval(str_query_orm)
                
            
            Auditoria(request.user, "Carga Datos CIO", "CargaDatos", id_carga).saveAudit()
        
        elif archivo == "N1":
            id_carga = GetId('RiesgoCargasisteman1', 'RCARN1', 'idcargan1').get_id()
            for j in range(2, m_row + 1):
                id_n1 = GetId('RiesgoCargasisteman1', 'IDN1', 'idn1').get_id()
                str_query_orm = "RiesgoCargasisteman1( idn1 = '"+str(id_n1)+"', idcargan1 = '"+str(id_carga)+"', estadoregistro='ACTIVO', fecharegistro='"+str(datetime.now())+"', fechamodificacion='"+str(datetime.now())+"',"
                for i in range(1, max_col + 1):
                    cell_obj_column = sheet_obj.cell(row = 1, column = i)                 
                    cell_obj = sheet_obj.cell(row = j, column = i)
                    datos_celda = cell_obj.value
                    if datos_celda is None:
                        datos_celda = "-"
                    column = cell_obj_column.value.lower()
                    str_query_orm += str(column)+"='"+str(datos_celda)+"',"
                str_query_orm+=").save()"
                eval(str_query_orm)
            
            Auditoria(request.user, "Carga Datos N1", "CargaDatos", id_carga).saveAudit()
        
        elif archivo == "Unifica":                        
            for j in range(2, m_row + 1):
                id_carga = id_generator()
                str_query_orm = "RiesgoUnifica( idregistro = '"+str(id_carga)+"',"
                for i in range(1, max_col + 1):
                    cell_obj_column = sheet_obj.cell(row = 1, column = i)                 
                    cell_obj = sheet_obj.cell(row = j, column = i)
                    datos_celda = cell_obj.value                    
                    
                    column = cell_obj_column.value.lower()
                    if (column == 'qhorasdetencion' or column == 'qdiasdetencion' or column == 'qktsperdida' or column =='qtmfperdida' or column == 'qlibrasperdida' or column == 'montoperdidakusd') and datos_celda is None:
                        datos_celda = 0
                    if column == "fecha" and datos_celda is None:
                        datos_celda = None
                    if column == "fecha" and datos_celda is not None:                                                                   
                        datos_celda = validaFecha(datos_celda)                    
                    if datos_celda is None:
                        datos_celda = "-"
                        
                    str_query_orm += str(column)+"='"+str(datos_celda)+"',"
                str_query_orm+=").save()"
                print(str_query_orm)
                eval(str_query_orm)
            
            Auditoria(request.user, "Carga Datos Unifica", "CargaDatos", id_carga).saveAudit()

        return JsonResponse(True, safe=False)

def validaFecha(datos_celda):   
    try:
        print("la fecha ", datos_celda)
        datos_celda = str(datos_celda.year)+'-'+str(datos_celda.month)+'-'+str(datos_celda.day)
        return datos_celda
    except:
        return str(datetime.now().year)+"-"+str(datetime.now().month)+"-"+str(datetime.now().day)

def datosDashboard(request):
    if request.method == "GET":
        datos = list(RiesgoUnifica.objects.all().values())
        return JsonResponse(datos, safe=False)

@login_required
def Dashboard(request):
    AreaDeEvento = list(RiesgoListas.objects.filter(tipo="AreaDeEvento").values())
    familias = list(QueryFamilia())
    subprocesos = list(RiesgoRbsfamilia.objects.values("subproceso").distinct())
    clasificacion = list(RiesgoListas.objects.filter(tipo="Clasificaciones").values())
    TipoFalla = list(RiesgoListas.objects.filter(tipo="TipoFalla").values())

    data = {
            'familias':familias, 
            'subprocesos':subprocesos,
            'clasificacion':clasificacion,
            'AreaDeEvento':AreaDeEvento,                                         
            'TipoFalla':TipoFalla,
            }
    if 'readcreate' in request.session:        
        return render(request, "appCarga/datosDashboard.html", data)
    else:
        return HttpResponseRedirect(reverse('inicio_dashboard'))

def editarDatosDashboard(request):
    if request.method == "POST":
        id_dashboard = request.POST.get('txt-dashboard')
        Fecha = request.POST.get("Fecha")
        Gerencia = request.POST.get("Gerencia")
        SuperIntendencia = request.POST.get("SuperIntendencia")
        UnidadOperacional = request.POST.get("UnidadOperacional")
        AreaDeEvento = request.POST.get("AreaDeEvento")
        Sector = request.POST.get("Sector")
        Lugar = request.POST.get("Lugar")
        TipoObjeto = request.POST.get("TipoObjeto")
        Evento = request.POST.get("Evento")
        NivelImpacto = request.POST.get("NivelImpacto")
        Comentarios = request.POST.get("Comentarios")
        QHorasDetencion = request.POST.get("QHorasDetencion")
        QDiasDetencion = request.POST.get("QDiasDetencion")
        QKtsPerdida = request.POST.get("QKtsPerdida")
        QTMFPerdida = request.POST.get("QTMFPerdida")
        QLibrasPerdida = request.POST.get("QLibrasPerdida")
        MontoPerdidaKUSD = request.POST.get("MontoPerdidaKUSD")
        IdRiesgoAsociado = request.POST.get("IdRiesgoAsociado")
        Familia = request.POST.get("Familia")
        SubProceso = request.POST.get("SubProceso")
        Clasificacion = request.POST.get("Clasificacion")
        TipoFalla = request.POST.get("TipoFalla")

        RiesgoUnifica.objects.filter(idregistro=id_dashboard).update(
            fecha = Fecha,
            gerencia = Gerencia,
            superintendencia = SuperIntendencia,
            unidadoperacional = UnidadOperacional,
            area = AreaDeEvento,
            sector = Sector,
            lugar = Lugar,
            tipoobjeto = TipoObjeto,
            evento = Evento,
            nivelimpacto = NivelImpacto,
            comentarios = Comentarios,
            qhorasdetencion = QHorasDetencion,
            qdiasdetencion = QDiasDetencion,
            qktsperdida = QKtsPerdida,
            qtmfperdida = QTMFPerdida,
            qlibrasperdida = QLibrasPerdida,
            montoperdidakusd = MontoPerdidaKUSD,
            idriesgoasociado = IdRiesgoAsociado,
            familia = Familia,
            subproceso = SubProceso,
            clasificacion = Clasificacion,
            tipofalla = TipoFalla,)
        return HttpResponseRedirect(reverse('datos_dashboard'))


